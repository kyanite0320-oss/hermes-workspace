#!/usr/bin/env python3
"""
策划配置表处理引擎 v1
流程：读飞书源表 → 按映射写目标Excel → 输出报告
"""
import json, urllib.request, os, sys, re
from pathlib import Path
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from formula_eval import is_formula, evaluate_source_value

# ============ 配置 ============
FEISHU_APP_ID = "cli_aa9cc2a3d1b8dbdf"
FEISHU_SECRET = "IjtBonENwQlSkyqvzU2thcqd8XmffCxk"
TARGET_DIR = "/mnt/c/workspace/project/excel"
MAPPING_FILE = "/mnt/c/AI/hermes/策划配置表工作流/策划配置表映射.xlsx"

# ============ 飞书API ============
def get_flybook_token():
    req = urllib.request.Request(
        "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
        data=json.dumps({"app_id": FEISHU_APP_ID, "app_secret": FEISHU_SECRET}).encode(),
        headers={"Content-Type": "application/json"}
    )
    return json.loads(urllib.request.urlopen(req).read())["tenant_access_token"]

def read_sheet(token, sheet_token, sheet_id, range_str):
    url = f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{sheet_token}/values/{sheet_id}!{range_str}"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    data = json.loads(urllib.request.urlopen(req).read())
    return data.get("data", {}).get("valueRange", {}).get("values", [])

# ============ 读取映射 ============
def read_mapping_from_excel():
    """从策划配置表映射.xlsx 的怪物映射 Sheet 读取映射规则"""
    wb = openpyxl.load_workbook(MAPPING_FILE, data_only=True)
    ws = wb["怪物映射"]
    
    mappings = []
    for row in ws.iter_rows(min_row=5, values_only=True):  # row 5 = header, data from 6
        src_table = str(row[1] or "").strip()
        src_sheet = str(row[2] or "").strip()
        src_col = str(row[3] or "").strip()
        target_file = str(row[4] or "").strip()
        target_sheet = str(row[5] or "").strip()
        target_col = str(row[6] or "").strip()
        rule_text = str(row[7] or "").strip()
        note_text = str(row[8] or "").strip()
            
        if src_col and target_file:
            mappings.append({
                "src_col": src_col,
                "target_file": target_file,
                "target_sheet": target_sheet,
                "target_col": target_col,
                "rule_text": rule_text,
                "note_text": note_text,
                "src_col_idx": None
            })
    
    return mappings, str(ws.cell(row=20, column=1).value or "").strip()

# ============ 读取源数据 ============
def read_source_data(mappings):
    """从飞书拉取怪物属性数据"""
    token = get_flybook_token()
    rows = read_sheet(token, "LFaAsr4mVhnm9ktZQ7XcN4wMnJc", "223cbf", "A1:AC227")
    
    if not rows:
        print("❌ 源数据为空")
        return []
    
    headers = rows[0]
    data_rows = rows[1:]
    
    # 找到每列对应的索引
    for m in mappings:
        for i, h in enumerate(headers):
            if h and h.strip() == m["src_col"]:
                m["src_col_idx"] = i
                break
    
    return [headers, data_rows]

# ============ 处理数据 ============
def parse_numeric(val):
    """提取数值，去掉单位"""
    if val is None or str(val).strip() == "":
        return None
    s = str(val).strip()
    # 去掉 "秒"、"像素"、"ms" 等单位
    s = re.sub(r'[秒像素毫ms]', '', s).strip()
    try:
        return float(s) if "." in s else int(s)
    except:
        return None

def process_monster(mappings, source_data):
    """处理怪物数据并写入目标Excel"""
    headers, data_rows = source_data
    changes = {"chapterxm.xlsx": 0, "skillxm.xlsx": 0, "new_rows": 0, "errors": []}
    
    # 按目标文件分组
    file_mappings = {}
    for m in mappings:
        if m["src_col_idx"] is None:
            continue
        f = m["target_file"]
        if f not in file_mappings:
            file_mappings[f] = []
        file_mappings[f].append(m)
    
    for file_name, f_mappings in file_mappings.items():
        file_path = os.path.join(TARGET_DIR, file_name)
        if not os.path.exists(file_path):
            changes["errors"].append(f"❌ 目标文件不存在: {file_path}")
            continue
        
        wb = openpyxl.load_workbook(file_path)
        
        # 按目标Sheet分组
        sheet_groups = {}
        for m in f_mappings:
            sn = m["target_sheet"]
            if sn not in sheet_groups:
                sheet_groups[sn] = []
            sheet_groups[sn].append(m)
        
        for sheet_name, s_mappings in sheet_groups.items():
            if sheet_name not in wb.sheetnames:
                changes["errors"].append(f"❌ Sheet不存在: {file_name}/{sheet_name}")
                continue
            
            ws = wb[sheet_name]
            
            # 找到目标列的索引 (based on row 7 which has internal names like "id", "HP", etc.)
            target_col_indices = {}
            row7 = []
            for cell in ws[7]:
                row7.append(str(cell.value or "").strip())
            
            for m in s_mappings:
                tcol = m["target_col"]
                # Handle multi-target like "moveSpeed|maxMoveSpeed"
                for col_name in tcol.split("|"):
                    col_name = col_name.strip()
                    for i, h in enumerate(row7):
                        if h == col_name:
                            if col_name not in target_col_indices:
                                target_col_indices[col_name] = i
                            break
            
            if sheet_name == "monsterTypexm":
                # 更新怪物数据
                for row_idx in range(9, ws.max_row + 1):  # data starts from row 9
                    data_id = str(ws.cell(row=row_idx, column=1).value or "").strip()
                    if not data_id or data_id == "id":
                        continue
                    
                    # Find matching source row by name
                    # The source_id is constructed as monXX based on order, but we match by name
                    row_modified = False
                    
                    # Iterate source data
                    for src_row in data_rows:
                        src_name = str(src_row[1] or "").strip() if len(src_row) > 1 else ""
                        target_name = str(ws.cell(row=row_idx, column=2).value or "").strip()
                        
                        if src_name != target_name:
                            continue
                        
                        # Apply mappings for this sheet
                        for m in s_mappings:
                            if m["src_col_idx"] is None:
                                continue
                            
                            src_val = src_row[m["src_col_idx"]] if m["src_col_idx"] < len(src_row) else ""
                            target_cols = m["target_col"].split("|")
                            
                            for tcol in target_cols:
                                tcol = tcol.strip()
                                if tcol in target_col_indices:
                                    ci = target_col_indices[tcol]
                                    old_val = ws.cell(row=row_idx, column=ci + 1).value
                                    
                                    # Determine value based on rule
                                    note = m["rule_text"] + " " + m["note_text"]
                                    val = src_val
                                    
                                    # 全局规则：公式算出数值再写入
                                    if isinstance(src_val, str) and is_formula(src_val):
                                        evaled = evaluate_source_value(src_val, data_rows, headers)
                                        if evaled is not None:
                                            val = int(evaled) if evaled == int(evaled) else evaled
                                            ws.cell(row=row_idx, column=ci + 1).value = val
                                            row_modified = True
                                            continue
                                    
                                    # Handle "转毫秒" rule
                                    if "毫秒" in note:
                                        num = parse_numeric(val)
                                        if num is not None:
                                            val = num * 1000 if num < 1000 else num
                                    
                                    # Write value (never copy formulas)
                                    if val is not None and str(val).strip():
                                        try:
                                            val_num = float(val) if "." in str(val) else int(val)
                                            ws.cell(row=row_idx, column=ci + 1).value = val_num
                                        except:
                                            ws.cell(row=row_idx, column=ci + 1).value = val
                                        
                                        if old_val != ws.cell(row=row_idx, column=ci + 1).value:
                                            row_modified = True
                        
                        break
                    
                    if row_modified:
                        changes["chapterxm.xlsx"] += 1
            
            elif sheet_name == "skillxm":
                monster_wb = openpyxl.load_workbook(os.path.join(TARGET_DIR, "chapterxm.xlsx"))
                monster_ws = monster_wb["monsterTypexm"]
                
                updates_made = 0
                for row_idx in range(9, ws.max_row + 1):
                    skill_id = str(ws.cell(row=row_idx, column=1).value or "").strip()
                    if not skill_id or skill_id == "id":
                        continue
                    
                    for monster_row_idx in range(9, monster_ws.max_row + 1):
                        monster_skill = str(monster_ws.cell(row=monster_row_idx, column=14).value or "").strip()
                        if monster_skill != skill_id:
                            continue
                        
                        monster_name = str(monster_ws.cell(row=monster_row_idx, column=2).value or "").strip()
                        
                        for src_row in data_rows:
                            src_name = str(src_row[1] or "").strip() if len(src_row) > 1 else ""
                            if src_name != monster_name:
                                continue
                            
                            # Apply skill mappings
                            for m in s_mappings:
                                if m["src_col_idx"] is None:
                                    continue
                                src_val = src_row[m["src_col_idx"]] if m["src_col_idx"] < len(src_row) else ""
                                if src_val is None:
                                    src_val = ""
                                target_cols = m["target_col"].split("|")
                                
                                for tcol in target_cols:
                                    tcol = tcol.strip()
                                    if tcol in target_col_indices:
                                        ci = target_col_indices[tcol]
                                        old_val = ws.cell(row=row_idx, column=ci + 1).value
                                        
                                        # 全局规则：不复制公式，只写数值
                                        if isinstance(src_val, str) and re.search(r'[=*/()ROUNDUP,]+', src_val):
                                            continue
                                        
                                        note = m["rule_text"] + " " + m["note_text"]
                                        val = src_val
                                        
                                        if "毫秒" in note:
                                            num = parse_numeric(val)
                                            if num is not None:
                                                val = num * 1000 if num < 1000 else num
                                        
                                        if val is not None and str(val).strip():
                                            # Write as numeric if possible
                                            try:
                                                val_num = float(val) if "." in str(val) else int(val)
                                                ws.cell(row=row_idx, column=ci + 1).value = val_num
                                            except:
                                                ws.cell(row=row_idx, column=ci + 1).value = str(val)
                                            
                                            new_val = ws.cell(row=row_idx, column=ci + 1).value
                                            changes["skillxm.xlsx"] += 1
                            break
                        break
        
        wb.save(file_path)
        
        # 全局规则：目标表不留公式，全部转数值
        wb2 = openpyxl.load_workbook(file_path, data_only=True)
        wb2.save(file_path)
        print(f"   [formula→value] {file_name}")
    
    return changes

# ============ 主流程 ============
def main():
    print("=" * 50)
    print("策划配置表处理引擎 v1")
    print("=" * 50)
    
    # Step 1: 读映射
    print("\n① 读取映射规则...")
    mappings, target_dir_note = read_mapping_from_excel()
    print(f"   找到 {len(mappings)} 条映射")
    
    # Step 2: 读源数据
    print("\n② 读取飞书源数据...")
    source_data = read_source_data(mappings)
    print(f"   源数据 {len(source_data[1])} 行")
    
    # Step 3: 处理
    print("\n③ 处理并写入目标Excel...")
    for m in mappings:
        src_idx = m["src_col_idx"]
        if src_idx is not None:
            print(f"   {m['src_col']} → {m['target_file']}/{m['target_sheet']}/{m['target_col']} (规则: {m['rule_text'] or 'direct'})")
        else:
            print(f"   ⚠️ {m['src_col']} → 未找到匹配的源列")
    
    changes = process_monster(mappings, source_data)
    
    # Step 4: 报告
    print(f"\n④ 处理报告")
    print(f"   chapterxm.xlsx 更新: {changes['chapterxm.xlsx']} 行")
    print(f"   skillxm.xlsx 更新: {changes['skillxm.xlsx']} 行")
    if changes["errors"]:
        for e in changes["errors"]:
            print(f"   ⚠️ {e}")
    
    print("\n✅ 完成")

if __name__ == "__main__":
    main()
